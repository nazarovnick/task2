from openpyxl.workbook import Workbook
from tempfile import NamedTemporaryFile
from rest_framework import mixins
from django.shortcuts import render
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticatedOrReadOnly
from rest_framework.viewsets import GenericViewSet
from django.http import HttpResponse
from rest_framework.views import APIView
import csv
from .permissions import HasTokenOrReadAndCreate
from .models import *
from .serializers import *


# Create your views here.
def index_page(request):

    comments = Comment.objects.all()

    context = {
                'comments' : comments,
               }

    return render(request = request, template_name = 'index.html', context = context)

def ApiIndexPage(request):
    return render(request = request, template_name = 'api.html')


class APIMixin(mixins.CreateModelMixin,
               mixins.RetrieveModelMixin,
               mixins.ListModelMixin,
               mixins.UpdateModelMixin,
               mixins.DestroyModelMixin,
               GenericViewSet):
    pass


class CountryViewSet(APIMixin):

    queryset = Country.objects.all()
    serializer_class = CountrySerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)
    authentication_classes = (TokenAuthentication,)


class DeveloperViewSet(APIMixin):

    queryset = Developer.objects.all()
    serializer_class = DeveloperSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)
    authentication_classes = (TokenAuthentication,)


class CarViewSet(APIMixin):

    queryset = Car.objects.all()
    serializer_class = CarSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)
    authentication_classes = (TokenAuthentication,)


class CommentViewSet(APIMixin):

    queryset = Comment.objects.all()
    serializer_class = CommentSerializer
    permission_classes = (HasTokenOrReadAndCreate,)
    authentication_classes = (TokenAuthentication,)


class GenerateExcelView(APIView):


        def get(self, request):

            field_1 = 'Текст комментария'
            field_2 = 'Дата добавления'
            field_3 = 'E-mail автора'
            field_4 = 'Автомобиль'
            field_5 = 'Производитель'
            field_6 = 'Страна'
            field_7 = 'Дата начала производства'
            field_8 = 'Дата окончания производства'


            filetype = request.GET.get('type')
            print(request.GET)
            if filetype == 'csv':

                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="all_data.csv"'
                writer = csv.DictWriter(response, fieldnames=[field_1,
                                                              field_2,
                                                              field_3,
                                                              field_4,
                                                              field_5,
                                                              field_6,
                                                              field_7,
                                                              field_8])

                writer.writeheader()

                data = Comment.objects.all()
                for line in data:
                    writer.writerow({ field_1: line.comment_text,
                                      field_2: line.date_created.strftime('%d.%m.%Y'),
                                      field_3: line.author_email,
                                      field_4: line.car_key.name,
                                      field_5: line.car_key.developer_key.name,
                                      field_6: line.car_key.developer_key.country_key.name,
                                      field_7: line.car_key.date_production_start.strftime('%d.%m.%Y'),
                                      field_8: line.car_key.date_production_stop.strftime('%d.%m.%Y')
                                     })

                return response

            elif filetype == 'xlsx':
                filename = 'all_data.xlsx'
                wb = Workbook()
                ws = wb.active
                ws.title = "Workbook"

                ws[f'A1'], ws[f'B1'], \
                ws[f'C1'], ws[f'D1'], \
                ws[f'E1'], ws[f'F1'], \
                ws[f'G1'], ws[f'H1'] = \
                field_1, field_2, \
                field_3, field_4, \
                field_5, field_6, \
                field_7, field_8

                data = Comment.objects.all()
                row_counter = 2
                for line in data:

                    ws[f'A{row_counter}'] = line.comment_text
                    ws[f'B{row_counter}'] = line.date_created.strftime('%d.%m.%Y')
                    ws[f'C{row_counter}'] = line.author_email
                    ws[f'D{row_counter}'] = line.car_key.name
                    ws[f'E{row_counter}'] = line.car_key.developer_key.name
                    ws[f'F{row_counter}'] = line.car_key.developer_key.country_key.name
                    ws[f'G{row_counter}'] = line.car_key.date_production_start.strftime('%d.%m.%Y')
                    ws[f'H{row_counter}'] = line.car_key.date_production_stop.strftime('%d.%m.%Y')

                    row_counter += 1

                with NamedTemporaryFile() as tmp:
                    wb.save(tmp.name)
                    tmp.seek(0)
                    stream = tmp.read()

                response = HttpResponse(content = stream, content_type = 'application/ms-excel')
                response["Content-Disposition"] = 'attachment; filename="' + filename + '"'
                return response
